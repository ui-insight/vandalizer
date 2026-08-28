"""Tests for the KB validation *results* export builder.

``build_kb_validation_results_export`` reshapes a persisted KB ValidationRun's
``result_snapshot`` into the export payload + per-query rows that the
``validation-runs/{run_uuid}/export`` endpoint serves as CSV / XLSX / JSON.
"""

from __future__ import annotations

import csv
import datetime
import io
from types import SimpleNamespace

from app.services.kb_validation_export import (
    EXPORT_FORMAT_TAG,
    RESULT_COLUMNS,
    RUN_META_CSV_COLUMNS,
    build_kb_validation_results_export,
    render_results_csv,
    render_results_xlsx,
)


def _make_kb():
    return SimpleNamespace(
        uuid="kb-1",
        title="NSF PAPPG",
        tags=["v24.1"],
        total_sources=3,
        total_chunks=120,
        resource_config={"seed_id": "kb-nsf-pappg"},
        rag_config_override={"k": 6},
    )


def _make_queries():
    return [
        SimpleNamespace(
            uuid="q-1",
            query="What is the deadline?",
            category="factual",
            expected_answer="30 days after award",
            expected_source_labels=["PAPPG Ch. 2"],
            external_id="ext-1",
        ),
        SimpleNamespace(
            uuid="q-2",
            query="Who signs the budget?",
            category="factual",
            expected_answer="The AOR",
            expected_source_labels=[],
            external_id=None,
        ),
    ]


def _make_vr():
    return SimpleNamespace(
        uuid="run-1234567890",
        score=82.5,
        score_breakdown={"raw_score": 88.0, "final_score": 82.5},
        created_at=datetime.datetime(2026, 8, 1, 12, 30, 0, tzinfo=datetime.timezone.utc),
        result_snapshot={
            "mode": "judge+baseline",
            "judge_model": "claude-x",
            "raw_score": 88.0,
            "num_test_queries": 3,
            "source_health": {"ratio": 1.0},
            "chunk_coverage": {"ratio": 0.9},
            "retrieval_precision": {
                "avg_precision": 0.75,
                "num_queries_judged": 2,
                "num_queries_baselined": 2,
                "avg_judge_score": 0.85,
                "avg_baseline_score": 0.4,
                "avg_lift": 0.45,
                "judge_variance": 0.03,
                "discrimination_summary": {"useful": 2, "redundant": 0, "failing": 0, "other": 0},
                "details": [
                    {
                        # Fully judged query, merged retrieval + judge fields.
                        "query": "What is the deadline?",
                        "query_uuid": "q-1",
                        "category": "factual",
                        "precision": 1.0,
                        "retrieved_sources": ["PAPPG Ch. 2", "PAPPG Ch. 7"],
                        "expected_sources": ["PAPPG Ch. 2"],
                        "answer_match": True,
                        "actual_answer": "The deadline is 30 days after award.",
                        "baseline_answer": "I am not sure.",
                        "judge": {
                            "score": 0.9,
                            "verdict": "PASS",
                            "confidence": 0.95,
                            "reasoning": "Matches the expected answer, cited Ch. 2.",
                            "missing_facts": [],
                            "hallucinated_facts": [],
                        },
                        "baseline_judge": {"score": 0.2, "verdict": "FAIL"},
                        "lift": 0.7,
                        "discrimination": "useful",
                    },
                    {
                        # Judged row keyed only by query text (legacy retrieval
                        # detail without a query_uuid) — expected answer must
                        # still join via the query string.
                        "query": "Who signs the budget?",
                        "precision": 0.5,
                        "retrieved_sources": ["Budget guide"],
                        "expected_sources": [],
                        "answer_match": None,
                        "actual_answer": "The AOR signs it.",
                        "judge": {
                            "score": 0.8,
                            "verdict": "PASS",
                            "confidence": 0.8,
                            "reasoning": "Correct, terse.",
                            "missing_facts": ["countersignature rule"],
                            "hallucinated_facts": [],
                        },
                        "lift": None,
                        "discrimination": "useful",
                    },
                    {
                        # Query deleted since the run; per-query error path.
                        "query": "Deleted question?",
                        "query_uuid": "q-gone",
                        "precision": 0.0,
                        "error": "retrieval blew up",
                    },
                ],
            },
        },
    )


def _build(vr=None):
    return build_kb_validation_results_export(
        kb=_make_kb(),
        vr=vr or _make_vr(),
        test_queries=_make_queries(),
        catalog_version="1.3.1",
        exported_by_user_id="user-1",
        exported_at="2026-08-03T00:00:00+00:00",
    )


def test_rows_map_judge_and_retrieval_fields():
    _, _, rows = _build()
    assert len(rows) == 3

    r = rows[0]
    assert r["question"] == "What is the deadline?"
    assert r["expected_answer"] == "30 days after award"
    assert r["actual_answer"] == "The deadline is 30 days after award."
    assert r["baseline_answer"] == "I am not sure."
    assert r["judge_score"] == 0.9
    assert r["judge_verdict"] == "PASS"
    assert r["judge_reasoning"] == "Matches the expected answer, cited Ch. 2."
    assert r["baseline_score"] == 0.2
    assert r["lift"] == 0.7
    assert r["discrimination"] == "useful"
    assert r["retrieved_sources"] == ["PAPPG Ch. 2", "PAPPG Ch. 7"]
    assert r["external_id"] == "ext-1"
    assert r["answer_match"] is True


def test_expected_answer_joins_by_query_text_when_uuid_missing():
    _, _, rows = _build()
    r = rows[1]
    assert r["expected_answer"] == "The AOR"
    assert r["query_uuid"] == "q-2"  # recovered from the joined query
    assert r["missing_facts"] == ["countersignature rule"]


def test_deleted_query_exports_blank_expected_answer_and_error():
    _, _, rows = _build()
    r = rows[2]
    assert r["expected_answer"] == ""
    assert r["judge_verdict"] == ""
    assert r["error"] == "retrieval blew up"


def test_run_meta_and_payload_shape():
    payload, run_meta, rows = _build()
    assert payload["format"] == EXPORT_FORMAT_TAG
    assert payload["results"] is rows
    assert payload["knowledge_base"]["title"] == "NSF PAPPG"
    assert run_meta["run_uuid"] == "run-1234567890"
    assert run_meta["judge_model"] == "claude-x"
    assert run_meta["mode"] == "judge+baseline"
    assert run_meta["run_score"] == 82.5
    assert run_meta["catalog_version"] == "1.3.1"
    assert run_meta["kb_seed_id"] == "kb-nsf-pappg"
    assert run_meta["avg_lift"] == 0.45
    assert run_meta["source_health_ratio"] == 1.0
    assert run_meta["rag_config_override_current"] == {"k": 6}
    assert run_meta["run_created_at"] == "2026-08-01T12:30:00+00:00"


def test_csv_repeats_run_meta_and_joins_lists():
    _, run_meta, rows = _build()
    text = render_results_csv(run_meta, rows)
    parsed = list(csv.reader(io.StringIO(text)))
    assert parsed[0] == RUN_META_CSV_COLUMNS + RESULT_COLUMNS
    assert len(parsed) == 4  # header + 3 queries
    first = dict(zip(parsed[0], parsed[1]))
    assert first["run_uuid"] == "run-1234567890"
    assert first["kb_title"] == "NSF PAPPG"
    assert first["retrieved_sources"] == "PAPPG Ch. 2; PAPPG Ch. 7"
    # Every row repeats the run metadata so multi-run CSVs concatenate cleanly.
    third = dict(zip(parsed[0], parsed[3]))
    assert third["run_uuid"] == "run-1234567890"
    assert third["error"] == "retrieval blew up"


def test_xlsx_round_trips_summary_and_results():
    from openpyxl import load_workbook

    _, run_meta, rows = _build()
    data = render_results_xlsx(run_meta, rows)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Summary", "Results"]

    summary = {r[0].value: r[1].value for r in wb["Summary"].iter_rows(min_row=2)}
    assert summary["run_uuid"] == "run-1234567890"
    assert summary["judge_model"] == "claude-x"

    results = wb["Results"]
    header = [c.value for c in results[1]]
    assert header == RESULT_COLUMNS
    body = list(results.iter_rows(min_row=2, values_only=True))
    assert len(body) == 3
    first = dict(zip(header, body[0]))
    assert first["question"] == "What is the deadline?"
    assert first["judge_score"] == 0.9


def test_xlsx_strips_illegal_control_characters():
    from openpyxl import load_workbook

    _, run_meta, rows = _build()
    rows[0]["actual_answer"] = "bad\x00control\x08chars kept text"
    data = render_results_xlsx(run_meta, rows)
    wb = load_workbook(io.BytesIO(data))
    results = wb["Results"]
    header = [c.value for c in results[1]]
    first = dict(zip(header, next(results.iter_rows(min_row=2, values_only=True))))
    assert first["actual_answer"] == "badcontrolchars kept text"


# ---------------------------------------------------------------------------
# A run's own record of what it scored against
# ---------------------------------------------------------------------------


def test_a_recorded_expected_answer_survives_the_query_being_deleted():
    """The failure this closes: the export re-joined the *live* test set for
    expected_answer and external_id, so pruning the set blanked those columns
    on runs that had already scored those questions. The run kept its score and
    the answer that was given, and silently lost the thing the score was
    measured against — which is what makes the export useful at all.
    """
    from app.services.kb_validation_export import build_kb_validation_results_export

    vr = SimpleNamespace(
        uuid="run-1",
        created_at=None,
        score=88.0,
        model="judge-model",
        run_type="full",
        result_snapshot={
            "retrieval_precision": {
                "details": [{
                    "query_uuid": "gone-1",
                    "query": "Who signs the subaward?",
                    "expected_answer": "The authorized organizational representative",
                    "external_id": "SUB-002",
                    "precision": 1.0,
                }],
            },
        },
    )

    # The query has since been deleted, so nothing to re-join against.
    _payload, _meta, rows = build_kb_validation_results_export(
        kb=SimpleNamespace(uuid="kb-1", title="KB", tags=[], total_sources=1, total_chunks=2),
        vr=vr,
        test_queries=[],
        catalog_version=None,
        exported_by_user_id="u1",
        exported_at="2026-08-20T00:00:00Z",
    )

    assert rows[0]["expected_answer"] == "The authorized organizational representative"
    assert rows[0]["external_id"] == "SUB-002"


def test_an_older_run_still_falls_back_to_the_live_test_set():
    """Runs recorded before the fields were persisted have nothing on the row,
    so the re-join has to stay."""
    from app.services.kb_validation_export import build_kb_validation_results_export

    vr = SimpleNamespace(
        uuid="run-2",
        created_at=None,
        score=70.0,
        model="judge-model",
        run_type="full",
        result_snapshot={
            "retrieval_precision": {
                "details": [{"query_uuid": "q-1", "query": "Old question?", "precision": 0.5}],
            },
        },
    )
    live = [SimpleNamespace(
        uuid="q-1", query="Old question?", expected_answer="From the live set",
        external_id="OLD-1", category="factual", expected_source_labels=[],
    )]

    _payload, _meta, rows = build_kb_validation_results_export(
        kb=SimpleNamespace(uuid="kb-1", title="KB", tags=[], total_sources=1, total_chunks=2),
        vr=vr,
        test_queries=live,
        catalog_version=None,
        exported_by_user_id="u1",
        exported_at="2026-08-20T00:00:00Z",
    )

    assert rows[0]["expected_answer"] == "From the live set"
    assert rows[0]["external_id"] == "OLD-1"


# ---------------------------------------------------------------------------
# The overall score is a composite; the export must say what it is made of so
# nobody reads it as the judge's answer accuracy.
# ---------------------------------------------------------------------------


def test_run_meta_explains_the_overall_score_for_an_older_run():
    # The fixture snapshot predates score_components — derive from its ratios.
    _, run_meta, _ = _build()
    assert "composite" in run_meta["run_score_meaning"]
    assert "Answer accuracy" in run_meta["avg_judge_score_meaning"]
    assert run_meta["score_formula"].startswith("overall = 40% × answer accuracy (judge)")
    comps = {c["key"]: c for c in run_meta["score_components"]}
    assert comps["judge"] == {
        "key": "judge", "label": "answer accuracy (judge)", "weight": 0.40, "value": 85.0,
    }
    assert comps["retrieval_precision"]["value"] == 75.0
    assert comps["source_health"]["value"] == 100.0
    assert comps["chunk_coverage"]["value"] == 90.0


def test_run_meta_prefers_the_components_the_run_persisted():
    vr = _make_vr()
    vr.result_snapshot["score_formula"] = "overall = persisted"
    vr.result_snapshot["score_components"] = [
        {"key": "judge", "label": "answer accuracy (judge)", "weight": 0.4, "value": 12.0},
    ]
    _, run_meta, _ = _build(vr=vr)
    assert run_meta["score_formula"] == "overall = persisted"
    assert run_meta["score_components"][0]["value"] == 12.0


def test_csv_carries_formula_and_answer_accuracy_beside_the_overall_score():
    _, run_meta, rows = _build()
    parsed = list(csv.reader(io.StringIO(render_results_csv(run_meta, rows))))
    first = dict(zip(parsed[0], parsed[1]))
    assert first["run_score"] == "82.5"
    assert first["avg_judge_score"] == "0.85"
    assert first["score_formula"].startswith("overall = 40% × answer accuracy (judge)")


def test_xlsx_summary_renders_components_readably():
    from openpyxl import load_workbook

    _, run_meta, rows = _build()
    wb = load_workbook(io.BytesIO(render_results_xlsx(run_meta, rows)))
    summary = {r[0].value: r[1].value for r in wb["Summary"].iter_rows(min_row=2)}
    assert summary["score_components"] == (
        "40% × answer accuracy (judge) = 85.0; 25% × retrieval precision = 75.0; "
        "20% × source health = 100.0; 15% × chunk coverage = 90.0"
    )
    assert summary["run_score_meaning"].startswith("Overall quality score")


def test_rows_carry_truncation_flags_and_default_false_for_older_runs():
    """Rows written before the flags existed export False (not missing), and
    rows that recorded a cut export it under both the storage and the
    generation flag, independently for the baseline."""
    vr = _make_vr()
    details = vr.result_snapshot["retrieval_precision"]["details"]
    details[0]["actual_answer_truncated"] = True
    details[0]["generation_truncated"] = False
    details[0]["baseline_generation_truncated"] = True
    _, _, rows = build_kb_validation_results_export(
        kb=_make_kb(), vr=vr, test_queries=_make_queries(), catalog_version=None,
        exported_by_user_id="u", exported_at="2026-08-26T00:00:00Z",
    )
    first = rows[0]
    assert first["actual_answer_truncated"] is True
    assert first["generation_truncated"] is False
    assert first["baseline_answer_truncated"] is False
    assert first["baseline_generation_truncated"] is True
    older = rows[1]
    for key in ("actual_answer_truncated", "generation_truncated",
                "baseline_answer_truncated", "baseline_generation_truncated"):
        assert key in RESULT_COLUMNS
        assert older[key] is False
