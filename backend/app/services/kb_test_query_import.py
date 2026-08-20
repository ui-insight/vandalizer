"""Parse bulk test-query imports (CSV / XLSX) into normalized row dicts.

Evaluators keep complete validation test sets in spreadsheets; this module
turns an uploaded file into rows the import endpoint can upsert as
``KBTestQuery`` records. Pure parsing — no DB access — so it unit-tests
without Beanie.
"""

import csv
import datetime
import io
import logging

logger = logging.getLogger(__name__)

MAX_IMPORT_ROWS = 500

# Canonical field -> accepted header spellings (normalized: lowercased,
# underscores/extra whitespace collapsed to single spaces).
_HEADER_ALIASES: dict[str, set[str]] = {
    "query": {"question", "query", "prompt", "test question", "test query"},
    "expected_answer": {
        "expected answer", "answer", "expected response", "canonical answer",
    },
    "category": {
        "category", "type", "question type", "question category",
        "category/type", "question category/type",
    },
    "expected_source_labels": {
        "source", "sources", "section", "source or section", "source/section",
        "expected sources", "source labels", "expected source labels",
    },
    "notes": {"notes", "note", "comments", "comment"},
    "external_id": {
        "id", "question id", "external id", "stable id", "stable question id",
        "qid",
    },
    "expected_answer_contains": {"expected answer contains", "answer contains"},
}


class TestQueryImportError(ValueError):
    """File-level problem the user can act on (bad format, missing columns)."""



# Straight and curly pairs — spreadsheets autocorrect the former into the latter.
_QUOTE_PAIRS = {'"': '"', "\u201c": "\u201d", "\u201d": "\u201d"}


def _unwrap_quoted_cell(raw: str) -> tuple[str, bool]:
    """Strip one matched pair of quotes wrapping the *whole* cell.

    Returns ``(text, was_quoted)``. Only whole-cell quoting is honoured, and
    only where the quotes actually reach us: ``csv.reader`` consumes CSV
    quoting long before this module sees the value, so for CSV this is a no-op
    and the semicolon rule below is the real escape hatch. In XLSX a typed
    quote character does survive, and there it means what the author intended.
    """
    text = raw.strip()
    if len(text) >= 2 and text[0] in _QUOTE_PAIRS and text[-1] == _QUOTE_PAIRS[text[0]]:
        return text[1:-1].strip(), True
    return text, False


def _names_one_known_source(text: str, known_source_names) -> bool:
    """True when the cell, taken whole, names a single source in the KB.

    Same substring rule the validation run scores with, so a cell this keeps
    intact is one that run can actually credit.
    """
    if not known_source_names:
        return False
    needle = text.lower()
    return any(needle in name.lower() for name in known_source_names if name)


def _split_source_labels(raw: str, known_source_names=()) -> list[str]:
    """Split a source-label cell into individual labels.

    Commas are the obvious separator, but real source names contain them —
    "Subpart D-iii — Monitoring, Reporting, Remedies & Closeout" is one
    source, not three. Splitting it blindly invents labels that match no
    source in the KB, and every invented label also inflates the denominator
    of the retrieval-precision score, so the KB gets marked down for a
    spreadsheet formatting artifact rather than a retrieval failure.

    Three things stop that, in order:

      * A cell quoted as a whole is one label — but see ``_unwrap_quoted_cell``:
        for CSV the quotes are gone before we run, so this only helps XLSX.
      * Semicolons, when the cell contains any, are the *only* separator,
        leaving commas inside each label intact. This is the escape hatch that
        works in every format, and the one the template and help text teach.
      * Failing both, a cell that already names a source in this KB is kept
        whole. This is what rescues the case the other two were meant to cover
        for authors who wrote a plain comma-bearing source name and never read
        the instructions — which is how the 2 CFR 200 set came to be split.

    A cell matching none of the three still splits on commas, so every file
    that imported correctly before still does.
    """
    if not raw or not raw.strip():
        return []

    text, was_quoted = _unwrap_quoted_cell(raw)
    if not text:
        return []
    if was_quoted:
        return [text]
    if ";" in text:
        return _unwrap_parts(text.split(";"))
    if _names_one_known_source(text, known_source_names):
        return [text]
    return _unwrap_parts(text.split(","))


def _unwrap_parts(parts) -> list[str]:
    """Trim each split part and drop quotes wrapping one whole part.

    ``"Monitoring, Reporting"; Subpart E`` is the semicolon form written by an
    author who also quoted for good measure; the quotes are redundant there but
    must not end up inside the label, where they would stop it matching the
    source name.
    """
    out = []
    for part in parts:
        text, _ = _unwrap_quoted_cell(part)
        if text:
            out.append(text)
    return out


def _normalize_header(value: str) -> str:
    return " ".join(value.replace("_", " ").strip().lower().split())


def _cell_to_str(value) -> str:
    """Coerce a spreadsheet cell to clean text.

    Excel stores bare numbers as floats, so an ID column of ``1, 2, 3`` reads
    back as ``1.0, 2.0, 3.0`` — strip the spurious decimal.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return str(value).strip()


def _raw_rows_from_csv(data: bytes) -> list[list[str]]:
    # utf-8-sig eats the BOM Excel prepends to exported CSVs; Windows-Excel
    # exports may instead be cp1252, so fall back rather than reject the file.
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("cp1252", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _raw_rows_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Test-query import: unreadable xlsx: %s", e)
        raise TestQueryImportError(
            "Couldn't read the Excel file. Re-save it as .xlsx (or export as CSV) and try again."
        )
    try:
        ws = wb.active
        if ws is None:
            return []
        return [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_test_query_import(
    filename: str, data: bytes, known_source_names=(),
) -> tuple[list[dict], list[dict]]:
    """Parse an uploaded CSV/XLSX into (rows, row_errors).

    Each row dict carries the ``KBTestQuery`` user fields: ``query``,
    ``expected_answer``, ``expected_answer_contains``, ``category``,
    ``expected_source_labels`` (list), ``notes``, ``external_id`` — absent
    columns yield ``None``/``[]``. ``row_errors`` entries are
    ``{"row": <1-based spreadsheet row>, "error": str}`` for rows that were
    skipped but shouldn't fail the whole import.

    Raises ``TestQueryImportError`` for file-level problems.
    """
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        raw = _raw_rows_from_csv(data)
    elif lowered.endswith(".xlsx"):
        raw = _raw_rows_from_xlsx(data)
    elif lowered.endswith(".xls"):
        raise TestQueryImportError(
            "Legacy .xls files aren't supported. Re-save the spreadsheet as .xlsx or .csv."
        )
    else:
        raise TestQueryImportError(
            "Unsupported file type. Upload a .csv or .xlsx file."
        )

    # Drop fully-empty leading rows so a title row of blanks doesn't become
    # the header. header_offset keeps reported row numbers matching the sheet.
    header_offset = 0
    while raw and not any(c.strip() for c in raw[0] if isinstance(c, str)):
        raw.pop(0)
        header_offset += 1
    if not raw:
        raise TestQueryImportError("The file is empty.")

    header = raw[0]
    columns: dict[int, str] = {}
    for idx, cell in enumerate(header):
        normalized = _normalize_header(str(cell or ""))
        for field, aliases in _HEADER_ALIASES.items():
            if normalized in aliases and field not in columns.values():
                columns[idx] = field
                break
    if "query" not in columns.values():
        raise TestQueryImportError(
            'No question column found. The header row must include a "Question" '
            '(or "Query") column. Download the template for the expected format.'
        )

    data_rows = raw[1:]
    if len(data_rows) > MAX_IMPORT_ROWS:
        raise TestQueryImportError(
            f"The file has {len(data_rows)} rows; the limit is {MAX_IMPORT_ROWS} "
            "per import. Split it into smaller files."
        )

    rows: list[dict] = []
    errors: list[dict] = []
    for i, raw_row in enumerate(data_rows):
        # 1-based sheet row: skipped blanks + header + this row.
        sheet_row = header_offset + 1 + i + 1
        values = {field: "" for field in _HEADER_ALIASES}
        for idx, field in columns.items():
            if idx < len(raw_row):
                values[field] = _cell_to_str(raw_row[idx])
        if not any(v for v in values.values()):
            continue  # blank spacer row
        if not values["query"]:
            errors.append({"row": sheet_row, "error": "Missing question"})
            continue
        labels = _split_source_labels(
            values["expected_source_labels"], known_source_names,
        )
        rows.append({
            "row": sheet_row,
            "query": values["query"],
            "expected_answer": values["expected_answer"] or None,
            "expected_answer_contains": values["expected_answer_contains"] or None,
            "category": values["category"].lower() or None,
            "expected_source_labels": labels,
            "notes": values["notes"] or None,
            "external_id": values["external_id"] or None,
        })
    return rows, errors
