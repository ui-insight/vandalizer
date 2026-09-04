"""Multi-format text extraction (PDF, DOCX, XLSX, HTML, etc.).

Ported from Flask app/utilities/document_readers.py.
All functions are synchronous — safe for Celery workers.
"""

import io
import logging
import re
from datetime import date, datetime, time
from typing import NoReturn

from markitdown import MarkItDown

from app.services import pdf_hidden_text

logger = logging.getLogger(__name__)


class DocumentReadError(RuntimeError):
    """Text extraction failed for a document.

    Raised instead of returning an error-message string as if it were the
    document's text — a returned placeholder is indistinguishable from real
    content downstream (it gets token-counted, chunked, and embedded), so a
    crashed reader must fail the document visibly instead.
    """


MIN_PDF_TEXT_LENGTH = 100
MAX_XLSX_COMMENT_LEN = 500

# Below this classifier confidence we don't trust the local fast path and
# fall through to the existing OCR-first flow — same "prefer accuracy over
# speed when unsure" posture as the rest of this module.
_PDF_INSPECTOR_MIN_CONFIDENCE = 0.8

# A rendered grayscale pixel this bright or brighter counts as blank paper.
# Real content — even faint anti-aliased text — pulls pixels well below this.
_BLANK_PAGE_INK_THRESHOLD = 250


# A table cell that is exactly the pandas/openpyxl NaN sentinel, and nothing
# else. Anchored to cell boundaries so real words survive: a blind
# str.replace("NaN", "") corrupted a PI surname (Nanjing, NaNoparticle) and
# any prose that happened to contain the letters.
_NAN_CELL_RE = re.compile(r"(?<=\|)(\s*)(?:NaN|nan|NAN)(\s*)(?=\|)")


def clean_markdown_nans(markdown_content: str) -> str:
    """Blank out NaN-only table cells in markdown content."""
    cleaned = _NAN_CELL_RE.sub(r"\1\2", markdown_content)

    lines = cleaned.split("\n")
    filtered_lines = []
    for line in lines:
        if "|" in line:
            cells = [cell.strip() for cell in line.split("|")[1:-1]]
            if any(cell and cell != "---" for cell in cells) or all(
                cell in ["---", ""] for cell in cells
            ):
                filtered_lines.append(line)
        else:
            filtered_lines.append(line)

    return "\n".join(filtered_lines)


def convert_to_markdown(doc_path: str, keep_data_uris: bool = True) -> str:
    """Convert a document to Markdown format using MarkItDown."""
    md = MarkItDown(enable_plugins=False)
    result = md.convert(doc_path, keep_data_uris=keep_data_uris)
    return clean_markdown_nans(result.text_content)


def extract_text_from_pdf(pdf_path: str, report: dict | None = None) -> str:
    """Extract text from a PDF using PyMuPDF. See _pymupdf_extract_with_pages for the page-aware variant.

    ``report``, when given, receives ``{"hidden_text_unchecked": True}`` if the
    hidden-text scrub could not inspect the file (see pdf_hidden_text).
    """
    text, _ = _pymupdf_extract_with_pages(pdf_path)
    scrubbed, _ = pdf_hidden_text.scrub_pdf(pdf_path, text, report=report)
    return scrubbed


def _pymupdf_extract_with_pages(pdf_path: str) -> tuple[str, list[dict]]:
    """Extract PDF text via PyMuPDF, returning text plus per-page char offsets.

    Markers are ``[{"char_offset": int, "kind": "page", "value": page_number}]``
    one per page of the source PDF. Used by ``extract_text_with_markers`` so
    chunks can be tagged with their source page for citations.

    PyMuPDF preserves reading order in multi-column layouts and exposes form
    field values that PyPDF2 misses (NIH biosketches, NSF Current & Pending
    forms, etc. are common research-admin uploads).
    """
    import pymupdf

    parts: list[str] = []
    markers: list[dict] = []
    cursor = 0

    try:
        doc = pymupdf.open(pdf_path)
    except pymupdf.FileNotFoundError as e:
        # PyMuPDF's FileNotFoundError is its own class (a RuntimeError), not
        # the builtin, so ``except FileNotFoundError`` upstream — the task
        # layer's "file vanished mid-processing, warn don't page" handler and
        # extract_text_from_file's — silently never matched it. Re-raise as
        # the builtin so a missing file looks like a missing file everywhere.
        raise FileNotFoundError(str(e)) from e

    with doc:
        for i, page in enumerate(doc, start=1):
            page_text = page.get_text("text")
            field_lines: list[str] = []
            for widget in page.widgets() or []:
                value = (widget.field_value or "").strip()
                if not value:
                    continue
                label = (widget.field_label or widget.field_name or "").strip()
                field_lines.append(f"- {label}: {value}" if label else f"- {value}")
            if field_lines:
                page_text = (page_text or "") + "\n[Form fields]\n" + "\n".join(field_lines)

            if not page_text:
                continue

            # Page marker points at the start of this page's text.
            markers.append({"char_offset": cursor, "kind": "page", "value": i})
            if parts:
                # The "\n" join we add below contributes one character.
                cursor += 1
            parts.append(page_text)
            cursor += len(page_text)

    return "\n".join(parts), markers


def ocr_extract_text_from_pdf(
    pdf_path: str, retries: int = 3, report: dict | None = None,
) -> str:
    """Extract text from a PDF using the configured OCR endpoint.

    The request/response contract depends on the configured provider — see
    ``app.services.ocr_client``. Falls back gracefully (returns "") if the OCR
    service is unavailable or misconfigured; callers then use PyMuPDF.

    ``report``, when given, is filled in with what the returned string cannot
    say — notably ``{"partial": True}`` when the converter only managed part of
    the document. Optional so existing callers are unaffected.
    """
    # OCR endpoint is stored in the database via admin config (SystemConfig)
    from app.services import ocr_client
    from app.tasks import get_sync_db
    from app.utils.encryption import decrypt_value
    db = get_sync_db()
    cfg = db.system_config.find_one({}) or {}
    ocr_endpoint = cfg.get("ocr_endpoint", "")
    raw_api_key = cfg.get("ocr_api_key", "")
    ocr_api_key = decrypt_value(raw_api_key)
    provider = ocr_client.normalize_provider(cfg.get("ocr_provider"))
    options = cfg.get("ocr_options") or {}
    use_async = bool(cfg.get("ocr_async"))
    timeout = float(cfg.get("ocr_timeout_seconds") or ocr_client.DEFAULT_OCR_TIMEOUT)

    if not ocr_endpoint:
        logger.warning("OCR_ENDPOINT not configured — skipping OCR for %s", pdf_path)
        return ""

    # If decrypt_value returned the raw 'enc:' ciphertext, CONFIG_ENCRYPTION_KEY
    # is missing or wrong in this process (commonly the Celery worker env).
    if ocr_api_key.startswith("enc:"):
        logger.error(
            "OCR api key could not be decrypted — CONFIG_ENCRYPTION_KEY missing "
            "or mismatched in this worker. Fix the env var and restart Celery."
        )
        return ""

    logger.info(
        "Extracting text with OCR: provider=%s endpoint=%s async=%s key_set=%s "
        "key_len=%d file=%s",
        provider, ocr_endpoint, use_async, bool(ocr_api_key), len(ocr_api_key), pdf_path,
    )

    headers = {}
    if ocr_api_key:
        headers["Authorization"] = f"Bearer {ocr_api_key}"

    import time as _time

    import httpx
    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            with httpx.Client(timeout=timeout) as client:
                return ocr_client.convert(
                    client,
                    pdf_path=pdf_path,
                    endpoint=ocr_endpoint,
                    headers=headers,
                    provider=provider,
                    options=options,
                    use_async=use_async,
                    report=report,
                )
        except ocr_client.OcrRequestError as e:
            last_error = e
            logger.warning(
                "OCR attempt %d failed against %s: %s — body: %s",
                attempt + 1, ocr_endpoint, e, e.body,
            )
        except Exception as e:
            last_error = e
            logger.warning("OCR attempt %d raised: %s", attempt + 1, e)
        if not ocr_client.is_retryable(last_error):
            # Nothing another attempt can fix — stop burning attempts and let
            # the caller degrade to PyMuPDF.
            break
        if attempt < retries - 1:
            wait = getattr(last_error, "retry_after", None)
            _time.sleep(wait if wait is not None else 2 ** attempt)

    # These three attempts span about 3 seconds, which is a network blip, not an
    # outage. When the failure looks transient, raise so the task layer's own
    # backoff (minutes, not seconds) gets a turn — returning "" here is what
    # made a brief OCR outage indistinguishable from a document that genuinely
    # has no text. See #633.
    if last_error is not None and ocr_client.is_retryable(last_error):
        logger.warning(
            "OCR unavailable after %d attempts for %s; deferring to task retry",
            retries, pdf_path,
        )
        raise ocr_client.OcrUnavailableError(
            f"OCR service did not respond after {retries} attempts: {last_error}",
            retry_after=getattr(last_error, "retry_after", None),
        )

    # A permanent failure (or no OCR configured at all) is a handled
    # degradation — the caller falls back to PyMuPDF — so log at warning, not
    # error, and don't page Sentry on every attempt-exhaustion.
    logger.warning("OCR failed after %d attempts for %s", retries, pdf_path)
    return ""


def _apply_percent_format(value: object, number_format: object) -> object:
    """Render a percent-formatted number the way the sheet shows it.

    Excel stores 47.5% as 0.475 and carries the "%" in the cell's display
    format, which the text extraction dropped — so a fringe rate the sheet
    shows as "47.5%" reached the model as "0.475", a different number by two
    orders of magnitude. Currency and thousands separators are deliberately
    NOT reconstructed: "$150,000" and "150000" are the same quantity, and
    stripping the symbols keeps the value machine-readable.
    """
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return value
    if not isinstance(number_format, str) or not _has_percent_operator(number_format):
        return value
    scaled = value * 100
    # 10 places, not 4: a rate displayed to more than four decimals was being
    # silently rounded, and a genuinely tiny one (1e-7) collapsed to "0%" —
    # a real value reaching the model as zero.
    text = f"{scaled:.10f}".rstrip("0").rstrip(".")
    return f"{text or '0'}%"


def _has_percent_operator(number_format: str) -> bool:
    """Whether a `%` in the format is Excel's scale-by-100 operator.

    A `%` inside quotes or after a backslash is a literal suffix: `0"%"` and
    `0.0" %"` mean "show a percent sign", and such cells already hold 47.5,
    not 0.475. Treating those as the operator multiplied them by 100 — the
    same two-orders-of-magnitude error this function exists to prevent, in
    the other direction.
    """
    in_quotes = False
    i = 0
    while i < len(number_format):
        ch = number_format[i]
        if ch == "\\":
            i += 2
            continue
        if ch == '"':
            in_quotes = not in_quotes
        elif ch == "%" and not in_quotes:
            return True
        i += 1
    return False


def _stringify_cell_value(value: object) -> str:
    """Render an openpyxl cell value as a plain display string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        rounded = round(value, 4)
        text = f"{rounded:.4f}".rstrip("0").rstrip(".")
        return text or "0"
    return str(value)


def _format_xlsx_cell(value: object) -> str:
    """Render an openpyxl cell value as a pipe-table-safe string."""
    text = _stringify_cell_value(value)
    if not text:
        return ""
    return text.replace("\\", "\\\\").replace("|", r"\|").replace("\n", " ").strip()


def _evaluate_xlsx_formulas(xlsx_path: str) -> dict[tuple[str, str], object]:
    """Compute formula results with the `formulas` library.

    Returns a {(sheet_name_upper, coordinate_upper): value} map. Empty on
    failure — callers fall back to formula text.
    """
    try:
        import formulas
    except ImportError:
        return {}

    try:
        model = formulas.ExcelModel().loads(xlsx_path).finish()
        solution = model.calculate()
    except Exception as e:
        logger.warning("formulas evaluation failed for %s: %s", xlsx_path, e)
        return {}

    result: dict[tuple[str, str], object] = {}
    for key, cell in solution.items():
        # Keys look like: "'[file.xlsx]SHEETNAME'!A1" or with ranges "...!A1:B2"
        try:
            sheet_part, coord = key.split("!", 1)
            if ":" in coord:
                continue  # skip range entries
            sheet_name = sheet_part.split("]", 1)[1].rstrip("'").upper()
        except Exception:
            continue

        value = cell.value if hasattr(cell, "value") else cell
        try:
            if hasattr(value, "tolist"):
                unwrapped = value.tolist()
                while isinstance(unwrapped, list) and len(unwrapped) == 1:
                    unwrapped = unwrapped[0]
                value = unwrapped
        except Exception:
            pass

        # `formulas` returns its own error sentinels for #DIV/0 etc.
        type_name = type(value).__name__
        if type_name in {"XlError", "Empty"}:
            continue

        result[(sheet_name, coord.upper())] = value

    return result


def extract_text_from_xlsx(xlsx_path: str) -> str:
    """Extract every visible and hidden sheet from an .xlsx workbook.

    Why this exists: MarkItDown (pandas under the hood) silently drops
    cells that have a formula but no cached calculated value. Workbooks
    saved by Google Sheets or headless LibreOffice commonly land in that
    state, so entire budget columns come through as blank. This walker
    uses openpyxl directly, prefers cached values, evaluates missing
    formulas with the `formulas` library when available, and falls back
    to formula text as a last resort.
    """
    import openpyxl

    try:
        wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        logger.warning(
            "openpyxl failed on %s (%s) — falling back to MarkItDown", xlsx_path, e
        )
        return convert_to_markdown(xlsx_path, keep_data_uris=False)

    # Only run the formula engine if we actually need it (any uncached formulas).
    needs_eval = False
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        for r in range(1, (ws_v.max_row or 0) + 1):
            for c in range(1, (ws_v.max_column or 0) + 1):
                if ws_v.cell(row=r, column=c).value is None:
                    fv = ws_f.cell(row=r, column=c).value
                    if isinstance(fv, str) and fv.startswith("="):
                        needs_eval = True
                        break
            if needs_eval:
                break
        if needs_eval:
            break

    computed = _evaluate_xlsx_formulas(xlsx_path) if needs_eval else {}

    out: list[str] = []

    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        max_row = ws_v.max_row or 0
        max_col = ws_v.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        sheet_key = sheet_name.upper()
        grid: list[list[object]] = []
        for r in range(1, max_row + 1):
            row: list[object] = []
            for c in range(1, max_col + 1):
                cell_v = ws_v.cell(row=r, column=c)
                cached = cell_v.value
                formula = ws_f.cell(row=r, column=c).value
                if cached is None and isinstance(formula, str) and formula.startswith("="):
                    coord = cell_v.coordinate.upper()
                    evaluated = computed.get((sheet_key, coord))
                    value = evaluated if evaluated is not None else formula
                else:
                    value = cached
                row.append(_apply_percent_format(value, cell_v.number_format))
            grid.append(row)

        # Spread a merged range's value across every cell it covers. openpyxl
        # stores it only in the top-left cell, so a "TOTAL DIRECT COSTS"
        # header merged across B2:E2 arrived as one labelled cell followed by
        # blanks — the columns under it read as unlabelled. The merge is still
        # listed in the extras below; this makes the grid itself readable.
        for rng in ws_v.merged_cells.ranges:
            anchor = grid[rng.min_row - 1][rng.min_col - 1] if (
                rng.min_row - 1 < len(grid) and rng.min_col - 1 < max_col
            ) else None
            if anchor in (None, ""):
                continue
            # Labels only. Spreading a merged NUMBER invents quantities: a
            # total of 485,000 merged across B10:E10 would render four times
            # and a model summing the row reads 1,940,000 — worse than the
            # unlabelled columns this fix exists to correct.
            if isinstance(anchor, (int, float)) and not isinstance(anchor, bool):
                continue
            for r in range(rng.min_row, min(rng.max_row, max_row) + 1):
                for c in range(rng.min_col, min(rng.max_col, max_col) + 1):
                    if grid[r - 1][c - 1] in (None, ""):
                        grid[r - 1][c - 1] = anchor

        kept_rows = [row for row in grid if any(v not in (None, "") for v in row)]
        if not kept_rows:
            continue

        keep_col = [
            any(row[c] not in (None, "") for row in kept_rows) for c in range(max_col)
        ]
        trimmed = [[row[c] for c, keep in enumerate(keep_col) if keep] for row in kept_rows]

        header = [f"## {sheet_name}"]
        if ws_v.sheet_state != "visible":
            header.append(f"_(sheet is {ws_v.sheet_state})_")
        out.append("\n".join(header))

        lines = ["| " + " | ".join(_format_xlsx_cell(v) for v in row) + " |" for row in trimmed]
        if len(lines) > 1:
            sep = "| " + " | ".join("---" for _ in trimmed[0]) + " |"
            lines.insert(1, sep)
        out.append("\n".join(lines))

        extras: list[str] = []
        merged = list(ws_v.merged_cells.ranges)
        if merged:
            extras.append(f"_Merged ranges: {', '.join(str(r) for r in merged)}_")

        hidden_rows = [
            r for r in range(1, max_row + 1) if ws_v.row_dimensions[r].hidden
        ]
        hidden_cols = [
            col for col, dim in ws_v.column_dimensions.items() if dim.hidden
        ]
        if hidden_rows:
            extras.append(f"_Hidden rows: {hidden_rows}_")
        if hidden_cols:
            extras.append(f"_Hidden columns: {hidden_cols}_")

        comments = []
        for row in ws_v.iter_rows():
            for cell in row:
                if cell.comment and cell.comment.text:
                    text = cell.comment.text.strip()
                    if len(text) > MAX_XLSX_COMMENT_LEN:
                        text = text[:MAX_XLSX_COMMENT_LEN] + "…"
                    comments.append((cell.coordinate, text))
        if comments:
            extras.append("_Cell comments:_")
            extras.extend(f"- {coord}: {text}" for coord, text in comments)

        if extras:
            out.append("\n".join(extras))

    defined = list(wb_formulas.defined_names)
    if defined:
        block = ["## Defined names"]
        for name in defined:
            try:
                value = wb_formulas.defined_names[name].value
            except Exception:
                value = "?"
            block.append(f"- {name}: {value}")
        out.append("\n".join(block))

    return "\n\n".join(out).strip()


def extract_sheet_json_from_xlsx(xlsx_path: str) -> dict:
    """Render an .xlsx workbook as JSON sheets for the document viewer.

    Mirrors extract_text_from_xlsx's evaluation strategy so the viewer
    and OCR agree on formula results: prefer Excel-cached values, then
    fall back to the `formulas` library, then to the formula text.
    """
    import openpyxl

    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)

    needs_eval = False
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]
        for r in range(1, (ws_v.max_row or 0) + 1):
            for c in range(1, (ws_v.max_column or 0) + 1):
                if ws_v.cell(row=r, column=c).value is None:
                    fv = ws_f.cell(row=r, column=c).value
                    if isinstance(fv, str) and fv.startswith("="):
                        needs_eval = True
                        break
            if needs_eval:
                break
        if needs_eval:
            break

    computed = _evaluate_xlsx_formulas(xlsx_path) if needs_eval else {}

    sheets: list[dict] = []
    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        max_row = ws_v.max_row or 0
        max_col = ws_v.max_column or 0
        if max_row == 0 or max_col == 0:
            sheets.append({"name": sheet_name, "headers": [], "rows": [], "hidden": ws_v.sheet_state != "visible"})
            continue

        sheet_key = sheet_name.upper()
        grid: list[list[str]] = []
        for r in range(1, max_row + 1):
            row: list[str] = []
            for c in range(1, max_col + 1):
                cell_v = ws_v.cell(row=r, column=c)
                cached = cell_v.value
                formula = ws_f.cell(row=r, column=c).value
                if cached is None and isinstance(formula, str) and formula.startswith("="):
                    coord = cell_v.coordinate.upper()
                    evaluated = computed.get((sheet_key, coord))
                    row.append(_stringify_cell_value(evaluated if evaluated is not None else formula))
                else:
                    row.append(_stringify_cell_value(cached))
            grid.append(row)

        headers = grid[0] if grid else []
        rows = grid[1:] if len(grid) > 1 else []
        sheets.append({
            "name": sheet_name,
            "headers": headers,
            "rows": rows,
            "hidden": ws_v.sheet_state != "visible",
        })

    return {"sheets": sheets}


def extract_sheet_json_from_csv(csv_path: str) -> dict:
    """Render a .csv as the same sheet JSON shape the viewer expects.

    Parsing moved server-side so the browser stops running a spreadsheet parser
    over untrusted uploads; a CSV needs nothing more than the stdlib to do it.
    Delimiter sniffing covers the semicolon and tab variants Excel emits under
    non-US locales, falling back to a comma when the sample is inconclusive.
    """
    import csv as _csv

    raw = _read_text_with_fallback(csv_path)
    sample = raw[:8192]
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except _csv.Error:
        dialect = _csv.excel

    grid = [list(row) for row in _csv.reader(io.StringIO(raw), dialect)]
    width = max((len(r) for r in grid), default=0)
    grid = [r + [""] * (width - len(r)) for r in grid]

    headers = grid[0] if grid else []
    rows = grid[1:] if len(grid) > 1 else []
    return {"sheets": [{
        "name": "Sheet1", "headers": headers, "rows": rows, "hidden": False,
    }]}


def extract_sheet_json_from_xls(xls_path: str) -> dict:
    """Render a legacy binary .xls as the same sheet JSON shape.

    xlrd reads .xls only — it dropped .xlsx years ago — which is why it is a
    much smaller thing to depend on than a general spreadsheet library, and why
    it runs here rather than in a reader's browser.

    .xls has no cached-vs-formula split to reconcile: the format stores
    computed values, so what is read is what Excel last calculated.
    """
    import xlrd

    book = xlrd.open_workbook(xls_path)
    sheets = []
    for ws in book.sheets():
        grid = [
            [_stringify_cell_value(ws.cell_value(r, c)) for c in range(ws.ncols)]
            for r in range(ws.nrows)
        ]
        headers = grid[0] if grid else []
        rows = grid[1:] if len(grid) > 1 else []
        sheets.append({
            "name": ws.name,
            "headers": headers,
            "rows": rows,
            # xlrd exposes visibility as 0 visible / 1 hidden / 2 very hidden.
            "hidden": getattr(ws, "visibility", 0) != 0,
        })
    return {"sheets": sheets}


def _read_text_with_fallback(path: str) -> str:
    """Decode a text file, tolerating the encodings spreadsheets arrive in."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    with open(path, encoding="utf-8", errors="replace") as f:
        return f.read()


_DOCX_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _docx_text_of(element) -> str:
    """Concatenate all <w:t> and <w:delText> descendants of a DOCX element."""
    parts = []
    for tag in ("t", "delText"):
        for node in element.iter(f"{{{_DOCX_W_NS}}}{tag}"):
            if node.text:
                parts.append(node.text)
    return "".join(parts).strip()


def read_docx_markdown(docx_path: str) -> str:
    """Body text of a .docx as markdown — the one reader for every path.

    Upload ingestion used pypandoc (with a silent, unlogged fallback to
    MarkItDown) while chat attachments used MarkItDown only, so the same file
    yielded different text depending on how it entered the system. Both now
    come through here. Note the fallback is not an edge case: the Docker
    image installs no pandoc binary, so on the supported deploy pypandoc
    raises OSError on every call and MarkItDown does all the work — which is
    why a *missing binary* logs at info. Any other pypandoc failure means a
    host that normally converts with pandoc just switched readers for this
    one document (different table/list rendering than its neighbors), which
    is worth a warning.

    ``--track-changes=accept`` is passed explicitly (insertions kept,
    deletions dropped from the body). That is pandoc's documented default,
    but the default was an assumption about whichever pandoc binary a host
    happens to have; struck-through text in a budget must not depend on it.
    Deleted text is not lost — ``extract_docx_extras`` reports it, labeled
    as deleted. MarkItDown's mammoth backend likewise accepts revisions
    (``w:ins`` content is read, ``w:del`` is in its ignored-elements list —
    pinned by test against the installed package).
    """
    try:
        import pypandoc

        body = pypandoc.convert_file(
            docx_path, "markdown", extra_args=["--track-changes=accept"],
        )
    except OSError as e:
        # "No pandoc was found" — the expected state on the Docker deploy.
        logger.info("pandoc not available for %s (%s); reading with MarkItDown", docx_path, e)
        body = None
    except Exception as e:
        logger.warning(
            "pypandoc failed to convert %s (%s); reading with MarkItDown",
            docx_path, e,
        )
        body = None
    if body is None:
        body = convert_to_markdown(docx_path, keep_data_uris=False)
    # Applied on BOTH branches: leaving image refs in only the MarkItDown
    # output would make the same docx read differently on Docker (no pandoc)
    # vs a pandoc host — the per-environment divergence this function exists
    # to kill.
    return remove_images_from_markdown(body)


def extract_docx_extras(docx_path: str) -> str:
    """Pull comments and tracked-change deletions from a .docx file.

    Returns a markdown block to append after the body, or "" if there's
    nothing notable. Research admins live in Word comments during
    proposal review, and pypandoc/MarkItDown both drop them silently.

    Deletions are reported because the body (read with tracked changes
    accepted) no longer contains them — a struck-through dollar figure is
    review history worth seeing, clearly labeled as deleted. Insertions are
    deliberately NOT listed: accepted insertions are already part of the
    body, and listing them again put every inserted figure into the context
    window twice — a "sum the personnel costs" prompt double-counted them.
    """
    import defusedxml.ElementTree as ET
    import zipfile

    try:
        zf = zipfile.ZipFile(docx_path)
    except (zipfile.BadZipFile, FileNotFoundError):
        return ""

    sections: list[str] = []

    with zf:
        names = set(zf.namelist())

        if "word/comments.xml" in names:
            try:
                tree = ET.fromstring(zf.read("word/comments.xml"))
            except ET.ParseError:
                logger.warning(
                    "Could not parse word/comments.xml in %s — comments will "
                    "be missing from the extracted text", docx_path,
                )
                tree = None
            if tree is not None:
                lines = []
                for c in tree.findall(f"{{{_DOCX_W_NS}}}comment"):
                    author = c.attrib.get(f"{{{_DOCX_W_NS}}}author", "Unknown")
                    date = c.attrib.get(f"{{{_DOCX_W_NS}}}date", "")
                    text = _docx_text_of(c)
                    if not text:
                        continue
                    header = f"- **{author}**"
                    if date:
                        header += f" ({date})"
                    lines.append(f"{header}: {text}")
                if lines:
                    sections.append("## Comments\n" + "\n".join(lines))

        if "word/document.xml" in names:
            try:
                doc_tree = ET.fromstring(zf.read("word/document.xml"))
            except ET.ParseError:
                logger.warning(
                    "Could not parse word/document.xml in %s — tracked-change "
                    "deletions will be missing from the extracted text",
                    docx_path,
                )
                doc_tree = None
            if doc_tree is not None:
                changes: list[str] = []
                for el in doc_tree.iter(f"{{{_DOCX_W_NS}}}del"):
                    text = _docx_text_of(el)
                    if not text:
                        continue
                    author = el.attrib.get(f"{{{_DOCX_W_NS}}}author", "Unknown")
                    date = el.attrib.get(f"{{{_DOCX_W_NS}}}date", "")
                    suffix = f" ({date})" if date else ""
                    changes.append(f"- **Deleted** by {author}{suffix}: {text}")
                if changes:
                    sections.append(
                        "## Tracked changes\n"
                        "_The following text was deleted in tracked changes "
                        "and is NOT part of the document body above._\n"
                        + "\n".join(changes)
                    )

    return "\n\n".join(sections)


def remove_images_from_markdown(markdown_text: str) -> str:
    """Remove all image references and their size attributes from markdown text."""
    text = re.sub(r"!\[([^\]]*)\]\([^)]*\)", "", markdown_text)
    text = re.sub(r"!\[([^\]]*)\]\[[^\]]*\]", "", text)
    text = re.sub(r'\{[^}]*(?:width|height)\s*=\s*"[^"]*"[^}]*\}', "", text)
    text = re.sub(r'\{[^{}]*="[^"]*"[^{}]*\}', "", text)
    text = re.sub(r"^\s*\[[^\]]+\]:\s*[^\s]+.*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"\n\s*\n\s*\n", "\n\n", text)
    text = re.sub(r"^\s+$", "", text, flags=re.MULTILINE)
    return text.strip()


def _xlsx_sheet_markers(text: str) -> list[dict]:
    """Recover per-sheet char offsets from extract_text_from_xlsx output.

    The xlsx extractor uses unique ``## {sheet_name}`` headers as section
    boundaries. Parsing them back out is cheaper than refactoring the whole
    builder to thread markers through every branch.
    """
    markers: list[dict] = []
    for m in re.finditer(r"(?m)^## (.+?)$", text):
        markers.append({"char_offset": m.start(), "kind": "sheet", "value": m.group(1).strip()})
    return markers


def _interpolate_page_markers(text: str, num_pages: int) -> list[dict]:
    """Approximate page boundaries by spreading them evenly across the text.

    Used when we know the page count (from PyMuPDF) but the text body came
    from a service that didn't preserve page structure (the OCR endpoint).
    Treats the text as uniformly dense — a rough heuristic, but good enough
    for "this answer is from somewhere around page 234" citations.

    Markers carry ``"approximate": True`` so consumers can tell an estimated
    boundary from a measured one (``_pymupdf_extract_with_pages`` emits the
    same shape without the flag). Anything that shows a page number to a user
    or a model should hedge accordingly — a confident citation off an
    interpolated offset is a fabricated one. Markers persisted before this
    flag existed have no key, and read as exact.
    """
    if num_pages <= 0 or not text:
        return []
    length = len(text)
    step = max(1, length // num_pages)
    return [
        {
            "char_offset": min(length - 1, i * step),
            "kind": "page",
            "value": i + 1,
            "approximate": True,
        }
        for i in range(num_pages)
    ]


def pdf_page_count(pdf_path: str) -> int:
    """Cheap page-count read via PyMuPDF. Returns 0 if it can't open the file.

    0 means "unknown", not "no pages" — callers that gate on a page count
    (sparse-text detection) must treat it as absent rather than as a real
    measurement, which is why the failure is logged at warning here.
    """
    try:
        import pymupdf
        with pymupdf.open(pdf_path) as doc:
            return doc.page_count
    except Exception as e:
        logger.warning(
            "Could not read PDF page count for %s (%s) — sparse-text "
            "detection is disabled for this document", pdf_path, e,
        )
        return 0


def pdf_has_ocrable_content(pdf_path: str) -> bool:
    """True if any page could plausibly yield real text via OCR.

    The OCR endpoint is a vision-LLM service; handed a blank page it
    fabricates plausible text rather than returning nothing, and that
    fabrication is long enough to pass the MIN_PDF_TEXT_LENGTH acceptance
    gate. So before OCR runs, prove there is something to read.

    A page counts as having content if it has a text layer, an embedded
    image (scanned page), or a filled form field. Pages with none of those
    are rasterized and tested for ink — rendering is the ground truth,
    since OCR consumes the rendered page: vector/outlined text (fonts
    converted to curves) shows up as dark pixels, while a decorative white
    background rectangle does not. A page that rasterizes to uniform white
    gives OCR nothing but an invitation to hallucinate.

    Fails open (returns True) if the file can't be opened or a page can't
    be rendered, so OCR still gets its chance on odd-but-valid PDFs.
    """
    import pymupdf

    try:
        with pymupdf.open(pdf_path) as doc:
            for page in doc:
                if (page.get_text("text") or "").strip():
                    return True
                if page.get_images(full=True):
                    return True
                if any(
                    (w.field_value or "").strip() for w in page.widgets() or []
                ):
                    return True
                pix = page.get_pixmap(colorspace=pymupdf.csGRAY, alpha=False)
                if pix.samples and min(pix.samples) < _BLANK_PAGE_INK_THRESHOLD:
                    return True
    except Exception as e:
        logger.warning(
            "Blank-page precheck could not inspect %s (%s) — assuming content",
            pdf_path, e,
        )
        return True
    return False


def _local_markdown_extract_from_pdf(pdf_path: str) -> tuple[str, list[dict]] | None:
    """Classify a PDF locally and, if it's confidently text-based with no
    pages flagged for OCR, extract structured Markdown locally — skipping
    the OCR round-trip entirely.

    Most research-admin PDFs (proposals, budgets, reports) are digitally
    native, not scanned, so this fast path is expected to fire for the
    majority of uploads: sub-5ms classification, no network call, and
    Markdown with real table/heading structure that flat OCR/PyMuPDF text
    doesn't have.

    Returns None — signalling "not a fit for the fast path" — for anything
    scanned, image-based, mixed, low classifier confidence, or any error, so
    the caller falls through to the existing OCR-first flow unchanged. This
    function never raises.
    """
    try:
        import pdf_inspector
    except ImportError:
        return None

    try:
        classification = pdf_inspector.classify_pdf(pdf_path)
    except Exception as e:
        logger.warning("pdf-inspector classification failed for %s: %s", pdf_path, e)
        return None

    if (
        classification.pdf_type != "text_based"
        or classification.confidence < _PDF_INSPECTOR_MIN_CONFIDENCE
        or classification.pages_needing_ocr
    ):
        return None

    try:
        result = pdf_inspector.extract_pages_markdown(pdf_path)
    except Exception as e:
        logger.warning("pdf-inspector extraction failed for %s: %s", pdf_path, e)
        return None

    parts: list[str] = []
    markers: list[dict] = []
    cursor = 0
    for page in result.pages:
        page_text = page.markdown or ""
        if not page_text:
            continue
        markers.append({"char_offset": cursor, "kind": "page", "value": page.page + 1})
        if parts:
            cursor += 1
        parts.append(page_text)
        cursor += len(page_text)

    text = "\n".join(parts)
    if len(text.strip()) < MIN_PDF_TEXT_LENGTH:
        return None

    logger.info(
        "pdf-inspector fast path: extracted %d chars from %s locally, "
        "skipping OCR (confidence=%.2f, tables_on_pages=%s)",
        len(text), pdf_path, classification.confidence, result.pages_with_tables,
    )
    return text, markers


def _extract_pdf_text_and_markers(
    file_path: str, report: dict | None = None,
) -> tuple[str, list[dict]]:
    """The one PDF path: read the text, then remove what the page hides.

    Every reader below — the local Markdown fast path, the OCR service,
    PyMuPDF — reports a PDF's text layer, and a text layer can carry words
    the page never shows (render mode 3, white on white, sub-point type).
    Those words are how a prompt injection reaches chat, extraction and
    Deep Analysis dressed as the document's own content, so they are cut
    here, at the single point every caller goes through, rather than
    defended against separately in each prompt downstream.
    """
    text, markers = _read_pdf_text_and_markers(file_path, report=report)
    return pdf_hidden_text.scrub_pdf(file_path, text, markers, report=report)


def _read_pdf_text_and_markers(
    file_path: str, report: dict | None = None,
) -> tuple[str, list[dict]]:
    """Extract a PDF's text and page markers with the best reader available."""
    # A local dict when the caller passed none: the partial-conversion signal
    # the OCR client records here decides below whether page markers can be
    # emitted at all, so it is needed even when no caller wants the report.
    if report is None:
        report = {}
    if not pdf_has_ocrable_content(file_path):
        logger.warning(
            "PDF %s rendered blank on every page — skipping OCR so the "
            "vision model can't fabricate text",
            file_path,
        )
        return "", []

    fast_path = _local_markdown_extract_from_pdf(file_path)
    if fast_path is not None:
        return fast_path

    # Prefer OCR text for accuracy. When OCR is used we lose true page
    # boundaries, so fall back to interpolating against PyMuPDF's page
    # count — approximate, but enough for "around page N" citations.
    from app.services import ocr_client

    try:
        ocr_text = ocr_extract_text_from_pdf(file_path, report=report)
    except ocr_client.OcrUnavailableError:
        # Deliberately not swallowed: a transient outage must reach the task
        # layer so the whole extraction is retried later, rather than being
        # degraded to whatever PyMuPDF can scrape off a scanned page now.
        raise
    except Exception as e:
        logger.warning("OCR raised, falling back to PyMuPDF: %s", e)
        ocr_text = ""
    if ocr_text and len(ocr_text.strip()) >= MIN_PDF_TEXT_LENGTH:
        # Interpolation spreads the *source PDF's* page count uniformly over
        # whatever text OCR returned. When the conversion was partial, the
        # text covers some unknown fraction of those pages, so every marker —
        # hedged or not — is systematically wrong (400 page numbers spread
        # over text from 30 pages). No page beats a wrong page: citations
        # fall back to the document title.
        if report.get("partial"):
            logger.warning(
                "OCR conversion of %s was partial — suppressing page markers, "
                "citations will carry no page numbers",
                file_path,
            )
            return ocr_text, []
        num_pages = pdf_page_count(file_path)
        return ocr_text, _interpolate_page_markers(ocr_text, num_pages)
    # Falling back means the partial OCR text is not what we return, so the
    # partial-conversion warning must not survive onto the PyMuPDF result.
    ocr_report_partial = bool(report.get("partial"))
    ocr_report_errors = list(report.get("errors") or [])
    report.pop("partial", None)
    report.pop("errors", None)
    # OCR unavailable / too little text — PyMuPDF gives us exact boundaries.
    # The PyMuPDF pass is a page-boundary refinement over the OCR text, not a
    # hard requirement. If it fails (corrupt PDF, or the source file was
    # removed between the OCR read and here — a document deleted mid-
    # processing), a short-but-valid OCR result still beats losing the
    # extraction and crashing the task.
    try:
        return _pymupdf_extract_with_pages(file_path)
    except Exception as e:
        if ocr_text and ocr_text.strip():
            logger.warning(
                "PyMuPDF page extraction failed for %s (%s); using OCR text",
                file_path, e,
            )
            if ocr_report_partial:
                # We are shipping the partial OCR text after all — and, as
                # above, page markers interpolated against the full PDF's
                # page count over partial text would all be wrong.
                report["partial"] = True
                report["errors"] = ocr_report_errors
                return ocr_text, []
            return ocr_text, _interpolate_page_markers(
                ocr_text, pdf_page_count(file_path)
            )
        raise


def extract_text_with_markers(
    file_path: str, file_extension: str, report: dict | None = None,
) -> tuple[str, list[dict]]:
    """Like extract_text_from_file, but also returns per-location char offsets.

    Markers are a list of ``{"char_offset": int, "kind": "page"|"sheet",
    "value": int|str}`` entries. Used by the chunker to attach page / sheet
    metadata to ChromaDB chunks so retrieval results can cite their source.

    Locations that can't preserve structure (DOCX text, plaintext, code
    files) return an empty marker list — chunks from those documents simply
    omit page metadata in citations.
    """
    ext = file_extension.lower().lstrip(".")

    if ext == "pdf":
        return _extract_pdf_text_and_markers(file_path, report=report)

    if ext == "xlsx":
        text = extract_text_from_xlsx(file_path)
        return text, _xlsx_sheet_markers(text)

    # Other formats (docx, txt, html, code) — extract as text, no markers.
    return extract_text_from_file(file_path, ext), []


# Sample sizes for the binary sniff below. The head and tail are checked so a
# binary with a clean text preamble (self-extracting archive, tar whose first
# member is text) can't sneak its tail through, and the byte-level sniff runs
# BEFORE the file is fully read so a 500 MB binary is refused without ever
# materializing it as a Python str in the worker.
_BINARY_SNIFF_BYTES = 1_000_000
# Below this, one stray control byte dominates the ratio (a 19-byte DOS text
# file with its historical \x1a EOF marker is 5.3% "junk"); the density test
# is only meaningful over a real sample. NUL/decode checks still apply.
_BINARY_SNIFF_MIN_LENGTH = 256

# Fraction of never-printable characters above which content is judged to be
# a binary, not text. Uniformly distributed bytes (compressed or encrypted
# data) land around 13%; structured binaries higher. Legacy text is at or
# near zero — the sets below deliberately exclude the C1 range cp1252 uses
# for curly quotes, dashes and €, and ESC, so quote-heavy prose and
# ANSI-colored logs/terminal captures can't trip it.
_BINARY_JUNK_THRESHOLD = 0.05

# Code points no text encoding prints: C0 controls minus real whitespace and
# ESC (see above), DEL, and the five code points cp1252 leaves undefined.
_JUNK_ORDINALS = (
    (set(range(32)) - {ord(c) for c in "\t\n\r\f\x0b"} - {0x1B})
    | {0x7F, 0x81, 0x8D, 0x8F, 0x90, 0x9D}
)
# For bytes.translate / str.translate — deleting junk and measuring the
# shrinkage counts occurrences in C instead of a per-character Python loop.
_JUNK_BYTES = bytes(sorted(_JUNK_ORDINALS - {0}))  # NUL handled separately
_JUNK_STR_TABLE = dict.fromkeys(_JUNK_ORDINALS, None)


def _byte_junk_fraction(sample: bytes) -> float:
    """Never-printable density of a byte sample, NUL excluded.

    NUL is excluded *here* because UTF-16 text is half NULs at the byte
    level; the per-decode check below rejects NULs that survive decoding.
    """
    if len(sample) < _BINARY_SNIFF_MIN_LENGTH:
        return 0.0
    kept = sample.translate(None, _JUNK_BYTES)
    return (len(sample) - len(kept)) / len(sample)


def _looks_like_binary(text: str) -> bool:
    """Whether a decoded string is a binary file wearing a text coat.

    A permissive codec's decode "succeeding" proves nothing (latin-1 maps
    every byte to a character). Two signals separate binaries from text: NUL
    characters (no text encoding decodes to them — a NUL-interleaved result
    means the wrong codec was used), and the density of characters no text
    encoding uses for content.
    """
    sample = text[:_BINARY_SNIFF_BYTES]
    if not sample:
        return False
    if "\x00" in sample:
        return True
    if len(sample) < _BINARY_SNIFF_MIN_LENGTH:
        return False
    junk = len(sample) - len(sample.translate(_JUNK_STR_TABLE))
    return junk / len(sample) > _BINARY_JUNK_THRESHOLD


def _refuse_binary(file_path: str, file_extension: str) -> "NoReturn":
    desc = f".{file_extension} file" if file_extension else "file"
    # The refusal is deliberate and user-actionable — log it here (the
    # DocumentReadError passthrough below skips the generic handler's log),
    # at warning rather than error so ordinary binary uploads don't page.
    logger.warning("Refusing to ingest %s as text — content looks binary", file_path)
    raise DocumentReadError(
        f"This {desc} does not appear to contain readable text — it may be "
        "a binary format this system cannot read. If it is a document, "
        "re-save it as PDF, DOCX, or plain text (UTF-8) and upload that."
    ) from None


def read_text_file(file_path: str, file_extension: str) -> str:
    """Read a plain-text file, refusing content that is not actually text.

    The one text reader for both the known-text extensions (txt/csv/code…)
    and the unknown-extension fallback, so a binary gets the same actionable
    refusal whatever it is named, instead of a raw codec error on one path
    and a latin-1 mojibake ingest on the other.

    Encodings are tried best-first, and a decode only counts when the result
    looks like text: BOM-less UTF-16LE ASCII is *valid UTF-8* (as
    NUL-interleaved junk), so "utf-8 decoded it" alone proves nothing — the
    NUL check sends it on to the utf-16 codec instead. utf-16 itself is only
    attempted when the byte-level NUL fraction carries the interleaved-text
    signature (~50% for ASCII), because the codec happily "decodes" arbitrary
    even-length binaries into CJK soup that would pass the density gate.
    cp1252 is tried before latin-1 so legacy memos keep their curly quotes
    and € instead of surviving as C1 mojibake.
    """
    with open(file_path, "rb") as f:
        head = f.read(_BINARY_SNIFF_BYTES)
        if _byte_junk_fraction(head) > _BINARY_JUNK_THRESHOLD:
            _refuse_binary(file_path, file_extension)
        raw = head + f.read()

    if len(raw) > 2 * _BINARY_SNIFF_BYTES and (
        _byte_junk_fraction(raw[-_BINARY_SNIFF_BYTES:]) > _BINARY_JUNK_THRESHOLD
    ):
        _refuse_binary(file_path, file_extension)

    nul_fraction = (raw.count(0) / len(raw)) if raw else 0.0
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp1252", "latin-1"):
        if encoding == "utf-16" and not (0.25 <= nul_fraction <= 0.60):
            continue
        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        if _looks_like_binary(text):
            continue
        return text
    _refuse_binary(file_path, file_extension)


def extract_text_from_file(file_path: str, file_extension: str) -> str:
    """Extract text from a file based on its extension.

    This is the primary entry point used by document_tasks.
    """
    file_extension = file_extension.lower().lstrip(".")

    try:
        if file_extension == "pdf":
            # Same flow — and same hidden-text scrub — as the markers
            # variant; only the page offsets are dropped here. A transient
            # OCR outage still propagates unwrapped through the
            # ``ConnectionError`` clause below so the task layer retries it
            # rather than seeing a ``DocumentReadError`` its ``autoretry_for``
            # never matches (VANDALIZER-BACKEND-1F).
            return _extract_pdf_text_and_markers(file_path)[0]

        elif file_extension in ("html", "htm"):
            return convert_to_markdown(file_path, keep_data_uris=False)

        elif file_extension in ("txt", "md", "csv", "json", "xml", "log"):
            # Through the gated reader rather than a strict utf-8 open: a
            # binary named report.txt used to fail with a raw codec error
            # while report.dat got the actionable refusal, and a legacy
            # cp1252 .txt failed outright instead of decoding.
            return read_text_file(file_path, file_extension)

        elif file_extension == "xlsx":
            return extract_text_from_xlsx(file_path)

        elif file_extension == "docx":
            # Same reader as upload ingestion (document_tasks) — a chat
            # attachment must not read differently from the uploaded copy.
            body = read_docx_markdown(file_path)
            extras = extract_docx_extras(file_path)
            return (body.rstrip() + "\n\n" + extras) if extras else body

        elif file_extension in ("doc", "xls", "pptx", "ppt"):
            return convert_to_markdown(file_path, keep_data_uris=False)

        elif file_extension in ("py", "js", "java", "cpp", "c", "h", "css", "sql"):
            return read_text_file(file_path, file_extension)

        else:
            try:
                return convert_to_markdown(file_path, keep_data_uris=False)
            except Exception:
                # Unknown extension MarkItDown refused. The gated reader is
                # the last resort: it decodes real text (any of the cascade's
                # encodings) and refuses binaries with an actionable message.
                # latin-1 used to be the terminal step here bare — it cannot
                # raise on any byte sequence, so any unrecognized *binary*
                # decoded "successfully", was stored as raw_text, chunked,
                # embedded, and answered from.
                return read_text_file(file_path, file_extension)

    except FileNotFoundError:
        # A missing source file (deleted mid-processing, retention sweep, or a
        # stale/relative path) is benign — there is nothing to extract. Return
        # empty text rather than a placeholder that would masquerade as content,
        # and log at warning so it doesn't page Sentry as a fault.
        logger.warning("Source file not found for extraction: %s", file_path)
        return ""

    except ConnectionError:
        # ``OcrUnavailableError`` (a ConnectionError subclass) re-raised above
        # must reach the task layer as itself, not as a DocumentReadError.
        raise

    except DocumentReadError:
        # Already carries a user-facing message (e.g. the binary-content
        # refusal above) — re-wrapping would just prefix it with a second
        # "Could not read this file:".
        raise

    except Exception as e:
        # Raise instead of returning an error string: a returned
        # "[Error extracting content: …]" is non-empty, so it sailed past the
        # empty-text guard, was token-counted, chunked, and embedded into
        # ChromaDB as the document's entire content — the document showed as
        # processed and chat answered confidently from a one-line error
        # message. A reader crash must fail the document visibly.
        logger.error("Error extracting text from %s: %s", file_path, e)
        raise DocumentReadError(
            f"Could not read this {file_extension or 'file'}: {e!s}"
        ) from e
