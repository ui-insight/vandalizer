#!/usr/bin/env python3

from pathlib import Path

import openpyxl
import pandas as pd


def save_excel_to_html(excel_file_path, html_file_path) -> None:
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheets = wb.sheetnames
    active_sheet = wb.active
    active_sheet_index = sheets.index(active_sheet.title)

    active_html_path = Path(html_file_path)
    output_dir_path = active_html_path.parent
    doc_name = active_html_path.stem

    output_dir_path.mkdir(parents=True, exist_ok=True)

    for i, sheet_name in enumerate(sheets):
        sheet = wb[sheet_name]
        max_col = sheet.max_column
        values = sheet.iter_rows(values_only=True)
        values = [[cell if cell is not None else "" for cell in row] for row in values]
        df = pd.DataFrame(values, columns=range(1, max_col + 1))
        html_string = df.to_html(
            index=False,
            header=False,
            classes=["table", "table-striped"],
            border=0,
        )

        # Add navigation links
        nav_links = ""
        if i > 0:
            prev_sheet = sheets[i - 1]
            prev_file = (
                f"{doc_name}-{prev_sheet}.html"
                if i - 1 != active_sheet_index
                else active_html_path.name
            )
            nav_links += f'<a href="{prev_file}">Previous Sheet</a> | '
        if i < len(sheets) - 1:
            next_sheet = sheets[i + 1]
            next_file = (
                f"{doc_name}-{next_sheet}.html"
                if i + 1 != active_sheet_index
                else active_html_path.name
            )
            nav_links += f'<a href="{next_file}">Next Sheet</a>'

        # HTML styles
        style = """
        <style>
        table { border-width: 1px !important; }
        </style>
        """

        # Combine navigation links, style, and table in the final HTML
        final_html = f"{nav_links}<br><br>{style}{html_string}"

        # Determine the file path
        if i == active_sheet_index:
            final_html_file_path = active_html_path
        else:
            final_html_file_path = output_dir_path / f"{doc_name}-{sheet_name}.html"

        with open(final_html_file_path, "w") as file:
            file.write(final_html)
