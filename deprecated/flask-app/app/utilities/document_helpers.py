#!/usr/bin/env python3


from pathlib import Path

import nltk
import openpyxl
import pandas as pd
import PyPDF2


def save_excel_to_html(excel_file_path, html_file_path) -> None:
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheets = wb.sheetnames
    active_sheet = wb.active
    active_index = sheets.index(active_sheet.title)

    active_html_path = Path(html_file_path)
    output_dir_path = active_html_path.parent
    doc_name = active_html_path.stem
    output_dir_path.mkdir(parents=True, exist_ok=True)

    # ——— Full-page & table styling ———
    base_css = """
    <style>
      body {
        background-color: #ffffff;
        font-family: Arial, sans-serif;
        margin: 1rem;
        margin: 30px;
      }
      .nav-links {
        margin: 1rem 0;
      }
      .nav-links a {
        color: #007bff;
        text-decoration: none;
        margin: 0 .5rem;
      }
      .nav-links a:hover {
        text-decoration: underline;
      }
      table {
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        box-shadow: 0 2px 6px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
      }
      td, th {
        padding: .75rem 1rem;
        border-bottom: 1px solid #e9ecef;
      }
      tbody tr:nth-of-type(odd) {
        background-color: #f8f9fa;
      }
      th {
        background-color: #343a40;
        color: #ffffff;
        text-align: left;
      }
    </style>
    """

    for i, sheet_name in enumerate(sheets):
        sheet = wb[sheet_name]
        max_col = sheet.max_column

        # read values and build DataFrame
        values = [
            [cell if cell is not None else "" for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        df = pd.DataFrame(values, columns=range(1, max_col + 1))

        # convert to HTML table
        table_html = df.to_html(
            index=False,
            header=False,
            classes=["table", "table-striped"],
            border=0,
        )

        # ——— navigation links ———
        links = []
        if i > 0:
            prev = sheets[i - 1]
            prev_file = (
                f"{doc_name}-{prev}.html"
                if (i - 1) != active_index
                else active_html_path.name
            )
            links.append(f'<a href="{prev_file}">← Previous</a>')
        if i < len(sheets) - 1:
            nxt = sheets[i + 1]
            next_file = (
                f"{doc_name}-{nxt}.html"
                if (i + 1) != active_index
                else active_html_path.name
            )
            links.append(f'<a href="{next_file}">Next →</a>')
        nav_html = f'<div class="nav-links">{" | ".join(links)}</div>'

        # ——— wrap into full HTML ———
        full_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>{doc_name} – {sheet_name}</title>
  {base_css}
</head>
<body>
  {nav_html}
  {table_html}
  {nav_html}
</body>
</html>
"""

        # decide output path
        if i == active_index:
            out_path = active_html_path
        else:
            out_path = output_dir_path / f"{doc_name}-{sheet_name}.html"

        out_path.write_text(full_html, encoding="utf-8")


def chunk_pdf(pdf_path):
    pdf_file = open(pdf_path, "rb")
    pdf_reader = PyPDF2.PdfReader(pdf_file)

    # Extract text
    pdf_text = ""
    for page in range(len(pdf_reader.pages)):
        page_text = pdf_reader.pages[page].extract_text()
        pdf_text += page_text

    # Split text into paragraphs
    paragraphs = pdf_text.split("\n\n")

    # Tokenize paragraphs into sentences
    sentences = [nltk.sent_tokenize(p) for p in paragraphs]
    sentences = [s for sents in sentences for s in sents]

    # Tokenize sentences into words
    tokenized_sentences = [nltk.word_tokenize(sent) for sent in sentences]

    # Extract n-grams (phrases) from sentences
    n = 3
    ngrams = [list(nltk.ngrams(sent, n)) for sent in tokenized_sentences]

    phrases = [p for ng in ngrams for p in ng]

    three_word_phrases = [" ".join(phrase) for phrase in phrases if len(phrase) == 3]
    return sentences + three_word_phrases
